from openpyxl import load_workbook
import json

wb = load_workbook("Danh mục.xlsx", data_only=True)
ws = wb.active

items = []
current = None

for row in ws.iter_rows(min_row=3, values_only=True):

    stt = row[0]
    ten = str(row[1]).strip() if row[1] else ""
    sohieu = str(row[2]).strip() if row[2] else ""
    tinhtrang = str(row[3]).strip() if row[3] else ""
    ghichu = str(row[4]).strip() if row[4] else ""

    if stt is not None:

        if current:
            items.append(current)

        current = {
            "stt": int(stt),
            "ten": ten,
            "sohieu": [],
            "tinhtrang": [],
            "ghichu": ghichu
        }

        if sohieu:
            current["sohieu"].append(sohieu)

        if tinhtrang:
            current["tinhtrang"].append(tinhtrang)

    else:

        if current is None:
            continue

        if ten:
            current["ten"] += "; " + ten

        if sohieu:
            current["sohieu"].append(sohieu)

        if tinhtrang:
            current["tinhtrang"].append(tinhtrang)

        if ghichu:
            if current["ghichu"]:
                current["ghichu"] += "\n" + ghichu
            else:
                current["ghichu"] = ghichu

if current:
    items.append(current)

with open("../data/museum.json", "w", encoding="utf-8") as f:
    json.dump(items, f, ensure_ascii=False, indent=2)

print("Đã tạo", len(items), "hiện vật")